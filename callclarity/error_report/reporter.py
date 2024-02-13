import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Alignment
from openpyxl.utils import get_column_letter

class IncoherenceReporter:
    
    def __init__(
        self, 
        df,
        path,
        text_col,
        label_col,
        sheet_name = "report",
        min_row = 2,
        min_col = 2,
        freeze_col_name = "id",
    ):
        self.df = df
        self.path = path
        self.sheet_name = sheet_name
        self.wb = self._init_workbook(self.df,self.sheet_name)
        self.ws = self._init_worksheet(self.wb, self.sheet_name)

        # col names
        self.text_col = text_col
        self.label_col = label_col
        self.label_recommended_col = f"{label_col}_recommended"
        self.label_suggested_col = f"{label_col}_suggested"

        # freeze attributes 
        self.freeze_col_name = freeze_col_name
        
        # initiate handling metadata
        self.column_to_index = {cell.value: cell.column for cell in self.ws[1][1:]}
        
        self.min_row = min_row
        self.max_row = self.ws.max_row
        
        self.min_col = min_col
        self.max_col = self.ws.max_column

    def _init_workbook(self, df, sheet_name):
        
        df.reset_index(drop = True).to_excel(self.path, sheet_name = sheet_name, index = True)
        wb = openpyxl.load_workbook(self.path)

        return wb

    def _init_worksheet(self, wb, sheet_name):
        
        ws = wb[sheet_name]
        return ws
        
    def freeze_panes(self):
        freeze_cell_coord  = self.ws.cell(self.min_row,self.column_to_index[self.freeze_col_name]).coordinate
        self.ws.freeze_panes = freeze_cell_coord
        
    def disable_gridlines(self):
        self.ws.sheet_view.showGridLines = False
        
    def get_col_cells(self, col_index, include_all = False):
        min_row = 1 if include_all else self.min_row
        col_cells = [self.ws.cell(row, col_index) for row in range(min_row,self.max_row + 1)]
        return col_cells
    
    def get_row_cells(self, row_index, include_all = False):
        min_col = 1 if include_all else self.min_col
        row_cells = [self.ws.cell(row_index, col) for row in range(min_col,self.max_col + 1)]
        return row_cells
    
    def get_max_col_width(self, col_index):
        return max([
            len(str(cell.value)) 
            for cell in self.ws[get_column_letter(col_index)]
        ])
        
    def apply_format(self):
        
        # --- general formatting ---
        dashed_bottom_border = openpyxl.styles.borders.Border(
            left=openpyxl.styles.Side(border_style=None), 
            right=openpyxl.styles.Side(border_style=None), 
            top=openpyxl.styles.Side(border_style=None), 
            bottom=openpyxl.styles.Side(style="dashed", color = "d3d3d3"),
        )
        
        top_alignment = Alignment(vertical='top')
        
        for cell in [cell for cells in self.ws[f"A1":f"{get_column_letter(self.max_col)}{self.max_row}"] for cell in cells]:
            
            cell.border = dashed_bottom_border
            cell.alignment = top_alignment

        
        # --- alignment ---
        center_cols = [
            "cluster_id",
            "id",
            "cluster_priority",
            "priority",
            "incoherence_severity",
            "recommendation_confidence",
        ]
        for col in center_cols:
            for cell in self.get_col_cells(self.column_to_index[col], include_all = True): 
                cell.alignment = Alignment(horizontal='center', vertical = "top")
    
        left_cols = [
            self.label_col,
            self.label_recommended_col,
            self.label_suggested_col,
        ]
        for col in left_cols:
            for cell in self.get_col_cells(self.column_to_index[col], include_all = True): 
                cell.alignment = Alignment(horizontal='left', vertical = "top")
                
        wrap_text_cols = [self.text_col]
        for col in wrap_text_cols:
            for cell in self.get_col_cells(self.column_to_index[col], include_all = False): 
                cell.alignment = Alignment(horizontal='left', vertical = "top", wrap_text=True)

        # --- width ---
        cols_to_width = {
            "cluster_id": 10,
            "cluster_priority":14,
            "id":7,
            "priority": 11,
            "incoherence_severity": 18,
            "recommendation_confidence": 25,
            self.text_col:40,
            self.label_col:35,
            self.label_recommended_col:35,
            self.label_suggested_col:35,
        }
        for col, col_width in cols_to_width.items():
            self.ws.column_dimensions[get_column_letter(self.column_to_index[col])].width = col_width
            
        # --- borders ---
        medium_top_border = openpyxl.styles.borders.Border(
            top= openpyxl.styles.Side(style="medium"), 
            bottom=openpyxl.styles.Side(style="dashed", color = "d3d3d3"),
        )
        thin_top_border = openpyxl.styles.borders.Border(
            top= openpyxl.styles.Side(style="thin"), 
            bottom=openpyxl.styles.Side(style="dashed", color = "d3d3d3"),
        )
         
        cluster_id_col = self.column_to_index["cluster_id"]
        cluster_priority_col = self.column_to_index["cluster_priority"]
        prev_id = None
        prev_priority = None
        for id_cell,priority_cell in zip(self.get_col_cells(cluster_id_col), self.get_col_cells(cluster_priority_col)):
            
            curr_id = id_cell.value
            curr_priority = id_cell.value
            
            if ((prev_id != None) and (curr_id != prev_id)):

                # for all cells in that row, change border
                current_row_index = id_cell.row

                if ((prev_priority <= 1) and (curr_priority <= 1)):
                    
                    # for top priority clusters, use thick lines
                    for cell in self.ws[current_row_index]:
                        cell.border = medium_top_border

                else:
                    
                    # otherwise, use thin lines
                    for cell in self.ws[current_row_index]:
                        cell.border = thin_top_border

            prev_id = curr_id
            prev_priority = curr_priority
            
        # --- value format ---
        perc_cols = [
            "incoherence_severity",
            "recommendation_confidence",
        ]
        for perc_col in perc_cols:
            index_col = self.column_to_index[perc_col]
            for cell in self.get_col_cells(index_col, include_all=False):
                cell.number_format = '0.00%'
            
        # --- fill color ---
        index_col = self.column_to_index["priority"]
        priority_to_color = {
            0:("3A0751","FFFFFF"),
            1:("D1193E","FFFFFF"),
            2:("FBA465","000000"),
            3:("F2C85B","000000"),
        }
        priority_to_color = {
            p:(PatternFill(fgColor=c[0],fill_type='solid'),Font(color=c[1])) 
            for p,c in priority_to_color.items()
        }
        for cell in self.get_col_cells(index_col, include_all=False):
            cell_value = cell.value
            if cell_value in priority_to_color:
                cell.fill = priority_to_color[cell_value][0]
                cell.font = priority_to_color[cell_value][1]
                
        # filling the color of fields to be filled
        offwhite_color = PatternFill(fgColor="FAF9F6",fill_type='solid')
        lightgrey_color = PatternFill(fgColor="DADADA",fill_type='solid')
        
        suggested_col_index = self.column_to_index[self.label_suggested_col]
        priority_col_index = self.column_to_index["priority"]
        
        suggested_cells = self.get_col_cells(suggested_col_index, include_all=False)
        priority_cells = self.get_col_cells(priority_col_index, include_all=False)
        
        for suggested_cell, priority_cell in zip(suggested_cells, priority_cells):
            
            if priority_cell.value != None:
                suggested_cell.fill = offwhite_color
            else:
                suggested_cell.fill = lightgrey_color


        # --- group columns ---
        self.ws.column_dimensions.group(
            "L", 
            get_column_letter(self.max_col),
            outline_level=1,
            hidden=True
        )
            
    def save(self):
        self.wb.save(self.path)
        self.wb.close()

    def report(self):
        self.freeze_panes()
        self.disable_gridlines()
        self.apply_format()
        self.save()